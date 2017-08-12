#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Kilauea_Project
@author: bruce.eo.thomas
"""

"""
Script for the GPS data for each year.
You will have to change the PATHS to data (excel or txt) !  
"""

import numpy as np
# Excel modul 
import xlrd
from openpyxl import load_workbook
# Name variables
from collections import defaultdict

nsites = 66

#--------------------------------------------------------------------------
# Open useful data.
#-------------------------------------------------------------------------- 

# Open Excel
path = "/gps/Bruce/Kilaueav2/data/coords"
access = xlrd.open_workbook(path+"/alldata.xlsx")
lendata = load_workbook(filename=path+"/alldata.xlsx", read_only=True)

# Open all sheets
sheets = access.sheet_names()

# Sheet 1 : 2009
s2009 = access.sheet_by_name(sheets[2])
lens2009 = lendata['2009use'].max_row
# Sheet 2 : 2011
s2011 = access.sheet_by_name(sheets[5])
lens2011 = lendata['2011use'].max_row
# Sheet 3 : 2017
s2017 = access.sheet_by_name(sheets[7])
lens2017 = lendata['2017use'].max_row


#--------------------------------------------------------------------------
# Convert in Python transposed matrix.
#-------------------------------------------------------------------------- 

"""
Exemple of a matrix per year
For one year
array([[ id site ...],
       [ x ...],
       [ y ...],
       [ z ...]])
"""

m2009 = np.zeros((4,lens2009))
for i in range(4):
    for j in range(lens2009):
        m2009[i][j] = s2009.cell_value(j,i)
        
m2011 = np.zeros((4,lens2011))
for i in range(4):
    for j in range(lens2011):
        m2011[i][j] = s2011.cell_value(j,i)

m2017 = np.zeros((4,lens2017))
for i in range(4):
    for j in range(lens2017):
        m2017[i][j] = s2017.cell_value(j,i)
       
mYears = [m2009, m2011, m2017]
len_mYears = len(mYears)
mYears_name = [2009, 2011, 2017]


#--------------------------------------------------------------------------
# Table per site (1 to 66) with coordinates for each year.
#--------------------------------------------------------------------------

"""
Exemple of a matrix for site evolution
For one site
array([[ year ...],
       [ x ...],
       [ y ...],
       [ z ...]])
"""

# Site evolution
se = defaultdict(int)

# Loop on each site
for site in range(1, nsites+1):
    # Create matrix changing name 
    se['se_%02d' % site] = np.empty((4,len_mYears))
    se['se_%02d' % site][:] = np.NAN   
    # Loop on each year of data
    for year in range(len_mYears):
        m = mYears[year]
        # Loop to find the site
        for i in range(len(m[0])):
            se['se_%02d' % site][0][year] = mYears_name[year]
            # Check if data for this site
            if site == m[0][i]:
                se['se_%02d' % site][1][year] = m[1][i]
                se['se_%02d' % site][2][year] = m[2][i]
                se['se_%02d' % site][3][year] = m[3][i]

"""
for key, value in sorted(se.items()):
    print (key, value)
"""


#--------------------------------------------------------------------------
# Table of displacements for each point. 
#--------------------------------------------------------------------------

"""
This is between first year of measure and last year of measure.
Not always the same years for each site, so there will be many matrix like that.

Exemple of a matrix between 2 years.
For 2 years
array([[ id site ...],
       [ dx ...],
       [ dy ...],
       [ dz ...]])
"""

# Define matrix of years 
m2009_2017 = np.empty((4,nsites))
m2009_2017[:] = np.NAN
m2011_2017 = np.empty((4,nsites))
m2011_2017[:] = np.NAN

# Loop on each site
for site in range(1, nsites+1):
    # First lien for id sites
    m2009_2017[0][site-1] = site
    m2011_2017[0][site-1] = site
    # Test to know which year of start 
    if se['se_%02d' % site][1][0] != 'nan' and se['se_%02d' % site][1][-1] != 'nan':
        # m2009_2017
        m2009_2017[1][site-1] = se['se_%02d' % site][1][-1] - se['se_%02d' % site][1][0]
        m2009_2017[2][site-1] = se['se_%02d' % site][2][-1] - se['se_%02d' % site][2][0]
        m2009_2017[3][site-1] = se['se_%02d' % site][3][-1] - se['se_%02d' % site][3][0]
    if se['se_%02d' % site][1][1] != 'nan' and se['se_%02d' % site][1][-1] != 'nan':
        # m2011_2017
        m2011_2017[1][site-1] = se['se_%02d' % site][1][-1] - se['se_%02d' % site][1][0]
        m2011_2017[2][site-1] = se['se_%02d' % site][2][-1] - se['se_%02d' % site][2][0]
        m2011_2017[3][site-1] = se['se_%02d' % site][3][-1] - se['se_%02d' % site][3][0] 
    # else, we let nan 


print(m2009_2017)
print(m2011_2017)



